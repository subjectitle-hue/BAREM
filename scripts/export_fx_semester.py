"""Faz 7 — İ sayfası dönem FX kurları (A satır 22–54: USD/EUR/çeyrek altın)."""
import json
import sys
from pathlib import Path

from openpyxl.utils import get_column_letter

sys.path.insert(0, str(Path(__file__).resolve().parent))
from excel_io import ASSETS, index_sheet, load_workbook, num

# İ! satır eşlemesi (Excel A! formülleri: AHn / İ!AJ94 vb.)
ROW_USD_OCAK = 94
ROW_USD_TEMMUZ = 96
ROW_EUR_OCAK = 107
ROW_EUR_TEMMUZ = 109
ROW_GOLD_QUARTER = 133


def main():
    wb = load_workbook(data_only=True, read_only=True)
    idx = index_sheet(wb)

    by_year = {}
    for c in range(5, (idx.max_column or 0) + 1):
        year = idx.cell(7, c).value
        if not isinstance(year, (int, float)):
            continue
        y = int(year)
        usd_o = num(idx.cell(ROW_USD_OCAK, c).value)
        usd_t = num(idx.cell(ROW_USD_TEMMUZ, c).value) or usd_o
        eur_o = num(idx.cell(ROW_EUR_OCAK, c).value)
        eur_t = num(idx.cell(ROW_EUR_TEMMUZ, c).value) or eur_o
        gold_q = num(idx.cell(ROW_GOLD_QUARTER, c).value)
        if not any([usd_o, eur_o, gold_q]):
            continue
        by_year[str(y)] = {
            "usdOcak": usd_o,
            "usdTemmuz": usd_t,
            "eurOcak": eur_o,
            "eurTemmuz": eur_t,
            "goldQuarter": gold_q,
            "column": get_column_letter(c),
        }

    payload = {"byYear": by_year}
    ASSETS.mkdir(parents=True, exist_ok=True)
    out = ASSETS / "fx_semester_rates.json"
    out.write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")
    print(f"wrote {out} years={len(by_year)}")
    wb.close()


if __name__ == "__main__":
    main()
