"""Export TS and M sheets from Mehmet Gürer.xlsx to Android assets JSON."""
import io
import json
from pathlib import Path

import msoffcrypto
import openpyxl

SOURCE = Path(r"C:\Users\MEMO\Downloads\Mehmet Gürer.xlsx")
PASSWORD = "erdal"
OUT_DIR = Path(__file__).resolve().parent.parent / "app" / "src" / "main" / "assets" / "memur"


def load_workbook():
    with SOURCE.open("rb") as f:
        o = msoffcrypto.OfficeFile(f)
        o.load_key(password=PASSWORD)
        b = io.BytesIO()
        o.decrypt(b)
        b.seek(0)
    return openpyxl.load_workbook(b, data_only=True, read_only=True)


def export_ts(wb):
    ws = wb["TS"]
    rows = []
    for r in range(3, (ws.max_row or 0) + 1):
        kolu = ws.cell(r, 2).value
        nitelik = ws.cell(r, 3).value
        kod_prefix = ws.cell(r, 4).value
        hizmet_sinifi = ws.cell(r, 5).value
        sira = ws.cell(r, 6).value
        kod = ws.cell(r, 7).value
        detay = ws.cell(r, 8).value
        if not kolu or not kod:
            continue
        rows.append({
            "hizmetKolu": str(kolu).strip(),
            "nitelik": str(nitelik).strip() if nitelik else "",
            "kodPrefix": str(kod_prefix).strip() if kod_prefix else "",
            "hizmetSinifi": str(hizmet_sinifi).strip() if hizmet_sinifi else "",
            "sira": int(sira) if isinstance(sira, (int, float)) else None,
            "kod": str(kod).strip(),
            "detay": str(detay).strip() if detay else "",
        })
    return rows


def export_m(wb):
    ws = wb["M"]
    rows = []
    for r in range(3, (ws.max_row or 0) + 1):
        hs = ws.cell(r, 1).value
        barem = ws.cell(r, 2).value
        unvan = ws.cell(r, 3).value
        detay = ws.cell(r, 4).value
        derece = ws.cell(r, 5).value
        if not hs or not unvan:
            continue
        rows.append({
            "hizmetSinifi": str(hs).strip(),
            "barem": int(barem) if isinstance(barem, (int, float)) else None,
            "unvan": str(unvan).strip(),
            "detay": str(detay).strip() if detay else "",
            "derece": int(derece) if isinstance(derece, (int, float)) else None,
        })
    return rows


def main():
    wb = load_workbook()
    ts = export_ts(wb)
    m = export_m(wb)
    wb.close()

    OUT_DIR.mkdir(parents=True, exist_ok=True)
    (OUT_DIR / "ts_meslek.json").write_text(
        json.dumps(ts, ensure_ascii=False, indent=0),
        encoding="utf-8",
    )
    (OUT_DIR / "m_kadro.json").write_text(
        json.dumps(m, ensure_ascii=False, indent=0),
        encoding="utf-8",
    )
    print(f"TS rows: {len(ts)}")
    print(f"M rows: {len(m)}")
    print(f"Written to {OUT_DIR}")


if __name__ == "__main__":
    main()
