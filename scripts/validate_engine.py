"""Validate salary engine logic against Excel cached BN/BM values."""
import io
from pathlib import Path

import msoffcrypto
import openpyxl
from openpyxl.utils import column_index_from_string

SOURCE = Path(r"C:\Users\MEMO\Downloads\Mehmet Gürer.xlsx")
PASSWORD = "erdal"


def load_h():
    with SOURCE.open("rb") as f:
        o = msoffcrypto.OfficeFile(f)
        o.load_key(password=PASSWORD)
        b = io.BytesIO()
        o.decrypt(b)
        b.seek(0)
    return openpyxl.load_workbook(b, data_only=True, read_only=True)["H"]


def r2(v):
    return round(v * 100) / 100


def compute_pay_lines(h, col, bl):
    memur_k = h.cell(4, col).value
    taban_k = h.cell(5, col).value
    yan_k = h.cell(6, col).value
    en_y = h.cell(8, col).value
    bn = {}
    bn[68] = r2(bl[68] * memur_k)
    bn[69] = r2(bl[69] * memur_k)
    bn[70] = r2(bl[70] * taban_k)
    bn[71] = r2(bl[71] * memur_k)
    bn[72] = r2(bl[72] * en_y / 100)
    bn[78] = r2(bl[78] * yan_k)
    bn[81] = r2(bl[81] * en_y / 100)
    return r2(sum(bn.values()))


def main():
    h = load_h()
    bm = column_index_from_string("BM")
    bn = column_index_from_string("BN")
    bl_col = column_index_from_string("BL")
    bl = {r: float(h.cell(r, bl_col).value or 0) for r in range(68, 104)}

    brut_bm = compute_pay_lines(h, bm, bl)
    brut_bn = compute_pay_lines(h, bn, bl)
    print("brut BM", brut_bm, "exp", h.cell(132, bm).value)
    print("brut BN", brut_bn, "exp", h.cell(132, bn).value)

    bm136 = r2(h.cell(133, bm).value * h.cell(18, bn).value / 100)
    bn136 = r2(h.cell(133, bn).value * h.cell(18, bn).value / 100)

    for m in range(12):
        brut = brut_bm if m < 6 else brut_bn
        prim = bm136 if m < 6 else bn136
        gv = h.cell(142 + m, bn).value
        dv = h.cell(154 + m, bn).value
        net = r2(brut - prim - gv - dv)
        print(f"m{m+1} net {net} exp {h.cell(166 + m, bn).value}")


if __name__ == "__main__":
    main()
