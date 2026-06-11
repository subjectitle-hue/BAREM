"""Export Excel engine data (İ, M, V, FX) for Android Memur calculator."""
import io
import json
from pathlib import Path

import msoffcrypto
import openpyxl
from openpyxl.utils import get_column_letter

SOURCE = Path(r"C:\Users\MEMO\Downloads\Mehmet Gürer.xlsx")
PASSWORD = "erdal"
OUT = Path(__file__).resolve().parent.parent / "app" / "src" / "main" / "assets" / "engine"


def load_workbook(data_only: bool = True):
    with SOURCE.open("rb") as f:
        o = msoffcrypto.OfficeFile(f)
        o.load_key(password=PASSWORD)
        b = io.BytesIO()
        o.decrypt(b)
        b.seek(0)
    return openpyxl.load_workbook(b, data_only=data_only, read_only=False)


def index_sheet(wb):
    for name in wb.sheetnames:
        if name in ("İ", "\u0130"):
            return wb[name]
    raise RuntimeError("İ sheet not found")


def export_index(ws):
    """İ!E7:AK93 — keys in B, period headers in row 7."""
    headers = {}
    for c in range(5, 38):
        year = ws.cell(7, c).value
        period = ws.cell(8, c).value
        if year is None:
            continue
        if isinstance(year, (int, float)):
            y = str(int(year))
            key = f"{y}-{period}" if period else y
        else:
            key = str(year).strip()
            if period:
                key = f"{key}-{period}"
        headers[c] = key

    entries = []
    for r in range(7, 94):
        label = ws.cell(r, 2).value
        if not label:
            continue
        label = str(label).strip()
        values = {}
        for c, pk in headers.items():
            v = ws.cell(r, c).value
            if v is None:
                continue
            if isinstance(v, (int, float)):
                values[pk] = float(v)
            else:
                values[pk] = str(v)
        entries.append({"label": label, "values": values})

    return {"headers": list(headers.values()), "entries": entries}


def export_m_kadro(ws):
    col = {}
    for c in range(1, 15):
        h = ws.cell(2, c).value
        if not h:
            continue
        name = str(h).strip().upper()
        if name == "TABAN":
            col["taban"] = c
        elif name == "KIDEM":
            col["kidem"] = c
        elif "GOSTERGE" in name or name == "EK GÖSTERGE":
            col["ek"] = c

    def num(cell):
        return float(cell) if isinstance(cell, (int, float)) else 0.0

    rows = []
    for r in range(3, (ws.max_row or 0) + 1):
        hs = ws.cell(r, 1).value
        unvan = ws.cell(r, 3).value
        if not hs or not unvan:
            continue
        taban_c = col.get("taban", 8)
        kidem_c = col.get("kidem", 9)
        taban_v = num(ws.cell(r, taban_c).value)
        kidem_v = num(ws.cell(r, kidem_c).value)
        rows.append({
            "hizmetSinifi": str(hs).strip(),
            "barem": int(ws.cell(r, 2).value) if isinstance(ws.cell(r, 2).value, (int, float)) else None,
            "unvan": str(unvan).strip(),
            "detay": str(ws.cell(r, 4).value).strip() if ws.cell(r, 4).value else "",
            "derece": int(ws.cell(r, 5).value) if isinstance(ws.cell(r, 5).value, (int, float)) else None,
            "taban": taban_v,
            "kidem": kidem_v,
            "ekGosterge": str(ws.cell(r, col.get("ek", 7)).value).strip() if ws.cell(r, col.get("ek", 7)).value else "",
        })
    return rows


def export_v_derece_kademe(ws, periods):
    """V!F:AK — row key in F (derece/kademe), years in row 2."""
    col_by_period = {}
    for c in range(6, 38):
        year = ws.cell(2, c).value
        if year is None:
            continue
        if isinstance(year, (int, float)):
            col_by_period[c] = str(int(year))
        else:
            col_by_period[c] = str(year).strip()

    result = {}
    for r in range(3, (ws.max_row or 0) + 1):
        key = ws.cell(r, 6).value
        if not key:
            continue
        key = str(key).strip()
        per_vals = {}
        for c, year in col_by_period.items():
            v = ws.cell(r, c).value
            if isinstance(v, (int, float)):
                per_vals[year] = float(v)
        if per_vals:
            result[key] = per_vals
    return result


def export_fx(ws):
    """Gold / USD rows ~94-106 from İ sheet."""
    fx = {}
    for r in range(90, 110):
        label = ws.cell(r, 2).value
        if not label:
            continue
        label = str(label).strip()
        values = {}
        for c in range(7, 38):
            year = ws.cell(7, c).value
            period = ws.cell(8, c).value
            v = ws.cell(r, c).value
            if year is None or not isinstance(v, (int, float)):
                continue
            if isinstance(year, (int, float)):
                y = str(int(year))
                pk = f"{y}-{period}" if period else y
            else:
                pk = str(year).strip()
                if period:
                    pk = f"{pk}-{period}"
            values[pk] = float(v)
        if values:
            fx[label] = values
    return fx


def export_h_meta(ws):
    """Period keys on H sheet row 3 (BN column = 2025-Temmuz style)."""
    periods = []
    for c in range(5, 80):
        year = ws.cell(2, c).value
        sem = ws.cell(3, c).value
        if year is None or sem is None:
            continue
        y = int(year) if isinstance(year, (int, float)) else str(year).strip()
        sem_s = str(sem).strip()
        key = f"{y}-{sem_s}" if isinstance(y, int) else f"{y}-{sem_s}"
        periods.append({
            "col": get_column_letter(c),
            "year": y,
            "semester": sem_s,
            "key": key,
        })
    return {"periods": periods, "activePeriod": "2025-1.012556"}


def main():
    wb = load_workbook(data_only=True)
    idx_ws = index_sheet(wb)
    index_data = export_index(idx_ws)
    fx_data = export_fx(idx_ws)
    m_data = export_m_kadro(wb["M"])
    v_data = export_v_derece_kademe(wb["V"], index_data["headers"])
    h_meta = export_h_meta(wb["H"])
    wb.close()

    OUT.mkdir(parents=True, exist_ok=True)
    (OUT / "index_table.json").write_text(
        json.dumps(index_data, ensure_ascii=False), encoding="utf-8"
    )
    (OUT / "m_kadro_full.json").write_text(
        json.dumps(m_data, ensure_ascii=False), encoding="utf-8"
    )
    (OUT / "v_derece_kademe.json").write_text(
        json.dumps(v_data, ensure_ascii=False), encoding="utf-8"
    )
    (OUT / "fx_rates.json").write_text(
        json.dumps(fx_data, ensure_ascii=False), encoding="utf-8"
    )
    (OUT / "h_meta.json").write_text(
        json.dumps(h_meta, ensure_ascii=False), encoding="utf-8"
    )
    print(f"index entries: {len(index_data['entries'])}")
    print(f"m kadro: {len(m_data)}")
    print(f"v keys: {len(v_data)}")
    print(f"written to {OUT}")


if __name__ == "__main__":
    main()
