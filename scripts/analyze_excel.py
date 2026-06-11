"""Analyze Mehmet Gürer.xlsx for EDMR porting report."""
import io
import json
import re
import sys
from collections import Counter, defaultdict
from pathlib import Path

import msoffcrypto
import openpyxl
from openpyxl.utils import get_column_letter

SOURCE = Path(r"C:\Users\MEMO\Downloads\Mehmet Gürer.xlsx")
PASSWORD = "erdal"
OUT = Path(__file__).resolve().parent.parent / "docs" / "excel_analysis_report.json"


def load_workbook():
    with SOURCE.open("rb") as f:
        office = msoffcrypto.OfficeFile(f)
        try:
            office.load_key(password=PASSWORD)
        except Exception as e:
            print(f"PASSWORD_FAIL: {e}", file=sys.stderr)
            office.load_key(password=None)
        decrypted = io.BytesIO()
        office.decrypt(decrypted)
        decrypted.seek(0)
    return openpyxl.load_workbook(decrypted, data_only=False, read_only=False, keep_vba=True)


def sheet_stats(ws):
    max_row = ws.max_row or 0
    max_col = ws.max_column or 0
    formula_cells = 0
    data_cells = 0
    sample_formulas = []
    ref_sheets = Counter()

    for row in ws.iter_rows(min_row=1, max_row=min(max_row, 5000), max_col=min(max_col, 200)):
        for cell in row:
            if cell.value is None:
                continue
            data_cells += 1
            if isinstance(cell.value, str) and cell.value.startswith("="):
                formula_cells += 1
                if len(sample_formulas) < 8:
                    sample_formulas.append(f"{cell.coordinate}: {cell.value[:120]}")
                for m in re.findall(r"'?([^'!]+)'?!", str(cell.value)):
                    if m != ws.title:
                        ref_sheets[m] += 1
            elif hasattr(cell, "data_type") and cell.data_type == "f":
                formula_cells += 1
                if len(sample_formulas) < 8 and cell.value:
                    sample_formulas.append(f"{cell.coordinate}: {str(cell.value)[:120]}")

    return {
        "title": ws.title,
        "max_row": max_row,
        "max_col": max_col,
        "approx_data_cells_scanned": data_cells,
        "formula_cells_scanned": formula_cells,
        "sample_formulas": sample_formulas,
        "external_sheet_refs_in_formulas": dict(ref_sheets.most_common(15)),
    }


def main():
    report = {
        "source": str(SOURCE),
        "size_mb": round(SOURCE.stat().st_size / (1024 * 1024), 2),
        "password_used": True,
        "errors": [],
        "workbook": {},
    }

    try:
        wb = load_workbook()
    except Exception as e:
        report["errors"].append(f"load_failed: {e}")
        OUT.parent.mkdir(parents=True, exist_ok=True)
        OUT.write_text(json.dumps(report, ensure_ascii=False, indent=2), encoding="utf-8")
        print(json.dumps(report, ensure_ascii=False, indent=2))
        return

    report["workbook"]["sheet_names"] = wb.sheetnames
    report["workbook"]["sheet_count"] = len(wb.sheetnames)
    report["workbook"]["has_vba"] = bool(getattr(wb, "vba_archive", None))

    if hasattr(wb, "defined_names") and wb.defined_names:
        names = []
        for key, dn in wb.defined_names.items():
            names.append({"name": key, "value": str(dn)[:200]})
        report["workbook"]["defined_names"] = names[:80]
        report["workbook"]["defined_names_count"] = len(names)

    sheets_detail = []
    for name in wb.sheetnames:
        try:
            ws = wb[name]
            sheets_detail.append(sheet_stats(ws))
        except Exception as e:
            sheets_detail.append({"title": name, "error": str(e)})

    report["sheets"] = sheets_detail

    # Data validations (first pass on active sheets)
    validations = []
    for name in wb.sheetnames[:30]:
        ws = wb[name]
        if hasattr(ws, "data_validations") and ws.data_validations:
            for dv in ws.data_validations.dataValidation:
                validations.append({
                    "sheet": name,
                    "type": dv.type,
                    "formula1": str(dv.formula1)[:150] if dv.formula1 else None,
                    "sqref": str(dv.sqref)[:100] if dv.sqref else None,
                })
    report["data_validations_sample"] = validations[:50]
    report["data_validations_count"] = len(validations)

    wb.close()

    OUT.parent.mkdir(parents=True, exist_ok=True)
    OUT.write_text(json.dumps(report, ensure_ascii=False, indent=2), encoding="utf-8")
    print(json.dumps(report, ensure_ascii=False, indent=2))


if __name__ == "__main__":
    main()
