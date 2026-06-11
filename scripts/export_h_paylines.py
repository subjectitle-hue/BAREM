"""Export H pay rows (68-103) formulas + cached BN values for Android engine."""
import io
import json
from pathlib import Path

import msoffcrypto
import openpyxl

SOURCE = Path(r"C:\Users\MEMO\Downloads\Mehmet Gürer.xlsx")
PASSWORD = "erdal"
OUT = Path(__file__).resolve().parent.parent / "app" / "src" / "main" / "assets" / "engine" / "h_paylines.json"


def load_wb(data_only: bool):
    with SOURCE.open("rb") as f:
        o = msoffcrypto.OfficeFile(f)
        o.load_key(password=PASSWORD)
        b = io.BytesIO()
        o.decrypt(b)
        b.seek(0)
    return openpyxl.load_workbook(b, data_only=data_only, read_only=True)


def main():
    wb_f = load_wb(False)
    wb_v = load_wb(True)
    h_f = wb_f["H"]
    h_v = wb_v["H"]
    from openpyxl.utils import column_index_from_string
    col_bl = column_index_from_string("BL")
    col_bn = column_index_from_string("BN")
    lines = []
    for r in range(68, 104):
        label = h_f.cell(r, 2).value
        if not label:
            continue
        lines.append({
            "row": r,
            "label": str(label).strip(),
            "bl": h_v.cell(r, col_bl).value,
            "bn_formula": str(h_f.cell(r, col_bn).value or ""),
            "bn_cached": float(h_v.cell(r, col_bn).value or 0),
        })
    for r in (178, 179, 180, 181):
        lines.append({
            "row": r,
            "label": str(h_f.cell(r, 2).value or f"row{r}").strip(),
            "bn_formula": str(h_f.cell(r, col_bn).value or ""),
            "bn_cached": float(h_v.cell(r, col_bn).value or 0),
        })
    OUT.parent.mkdir(parents=True, exist_ok=True)
    OUT.write_text(json.dumps({"periodColumn": "BN", "lines": lines}, ensure_ascii=False, indent=2), encoding="utf-8")
    print("wrote", OUT, "lines", len(lines))


if __name__ == "__main__":
    main()
