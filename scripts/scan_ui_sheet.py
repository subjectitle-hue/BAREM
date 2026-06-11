import io
import json
from pathlib import Path

import msoffcrypto
import openpyxl

SOURCE = Path(r"C:\Users\MEMO\Downloads\Mehmet Gürer.xlsx")
PASSWORD = "erdal"


def load():
    with SOURCE.open("rb") as f:
        o = msoffcrypto.OfficeFile(f)
        o.load_key(password=PASSWORD)
        b = io.BytesIO()
        o.decrypt(b)
        b.seek(0)
    return openpyxl.load_workbook(b, data_only=False, read_only=False)


def dump(ws, r1=1, r2=80, c1=1, c2=20):
    rows = []
    for r in range(r1, r2 + 1):
        row = []
        for c in range(c1, c2 + 1):
            cell = ws.cell(r, c)
            v = cell.value
            if v is None:
                row.append(None)
            else:
                s = str(v)
                if s.startswith("="):
                    row.append({"f": s[:180]})
                else:
                    row.append(s[:100])
        if any(x is not None for x in row):
            rows.append({"r": r, "cells": row})
    return rows


def main():
    wb = load()
    print("names:", [(i, repr(n)) for i, n in enumerate(wb.sheetnames)])
    # UI sheet likely index 4 (İ)
    for name in wb.sheetnames:
        if name in ("G", "M", "V"):
            continue
        ws = wb[name]
        print("\n####", repr(name), ws.max_row, ws.max_column)
        data = dump(ws, 1, min(120, ws.max_row or 1), 1, min(25, ws.max_column or 1))
        for block in data[:60]:
            print(block)
    wb.close()


if __name__ == "__main__":
    main()
