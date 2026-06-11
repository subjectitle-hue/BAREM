"""Export TS meslek kod -> tazminat / yan / ek gösterge (2025 sütunu)."""
import io
import json
from pathlib import Path

import msoffcrypto
import openpyxl

SOURCE = Path(r"C:\Users\MEMO\Downloads\Mehmet Gürer.xlsx")
OUT = Path(__file__).resolve().parent.parent / "app" / "src" / "main" / "assets" / "engine" / "ts_meslek_oran.json"
COL_2025 = 23


def load_wb():
    with SOURCE.open("rb") as f:
        o = msoffcrypto.OfficeFile(f)
        o.load_key(password="erdal")
        b = io.BytesIO()
        o.decrypt(b)
        b.seek(0)
    return openpyxl.load_workbook(b, data_only=True, read_only=True)


def main():
    wb = load_wb()
    ts = wb["TS"]
    by_kod: dict[str, dict] = {}
    for r in range(3, (ts.max_row or 0) + 1):
        kod = ts.cell(r, 7).value
        nitelik = ts.cell(r, 3).value
        if not kod or not nitelik:
            continue
        kod = str(kod).strip()
        nitelik = str(nitelik).strip().lower()
        val = ts.cell(r, COL_2025).value
        if not isinstance(val, (int, float)):
            continue
        entry = by_kod.setdefault(kod, {"tazminat": 0.0, "yanOdeme": 500.0, "ekOdeme": 90.0})
        if "tazminat" in nitelik:
            entry["tazminat"] = float(val)
        elif "ek" in nitelik and "ödeme" in nitelik.replace("ö", "o"):
            entry["ekOdeme"] = float(val)
        elif "yan" in nitelik:
            entry["yanOdeme"] = float(val)
        elif "ücret" in nitelik or "ucret" in nitelik:
            entry["yanOdeme"] = float(val)
    OUT.parent.mkdir(parents=True, exist_ok=True)
    OUT.write_text(json.dumps(by_kod, ensure_ascii=False, indent=2), encoding="utf-8")
    print("wrote", len(by_kod), "kod entries")


if __name__ == "__main__":
    main()
