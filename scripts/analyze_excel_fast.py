"""Fast pass: sheet dimensions + row counts via read_only."""
import io
import json
from pathlib import Path

import msoffcrypto
import openpyxl

SOURCE = Path(r"C:\Users\MEMO\Downloads\Mehmet Gürer.xlsx")
PASSWORD = "erdal"
OUT = Path(__file__).resolve().parent.parent / "docs" / "excel_sheets_overview.json"


def decrypt_buffer():
    with SOURCE.open("rb") as f:
        office = msoffcrypto.OfficeFile(f)
        office.load_key(password=PASSWORD)
        buf = io.BytesIO()
        office.decrypt(buf)
        buf.seek(0)
    return buf


def main():
    buf = decrypt_buffer()
    wb = openpyxl.load_workbook(buf, read_only=True, data_only=True, keep_links=False)
    sheets = []
    for name in wb.sheetnames:
        ws = wb[name]
        sheets.append({
            "title": name,
            "max_row": ws.max_row,
            "max_column": ws.max_column,
        })
    wb.close()
    data = {"sheet_count": len(sheets), "sheets": sheets}
    OUT.parent.mkdir(parents=True, exist_ok=True)
    OUT.write_text(json.dumps(data, ensure_ascii=False, indent=2), encoding="utf-8")
    print(json.dumps(data, ensure_ascii=False, indent=2))


if __name__ == "__main__":
    main()
