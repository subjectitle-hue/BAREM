import io
import json
from pathlib import Path

import msoffcrypto
import openpyxl

SOURCE = Path(r"C:\Users\MEMO\Downloads\Mehmet Gürer.xlsx")
PASSWORD = "erdal"


def load():
    with SOURCE.open("rb") as f:
        o = msoffcrypto.OfficeFile(f)
        o.load_key(password=PASSWORD)
        b = io.BytesIO()
        o.decrypt(b)
        b.seek(0)
    return openpyxl.load_workbook(b, read_only=True, data_only=True)


def sample_sheet(ws, rows=25, cols=12):
    grid = []
    for r in range(1, min(rows + 1, (ws.max_row or 0) + 1)):
        row = []
        for c in range(1, min(cols + 1, (ws.max_column or 0) + 1)):
            v = ws.cell(r, c).value
            if v is not None:
                s = str(v)
                row.append(s[:80] if len(s) > 80 else s)
            else:
                row.append(None)
        grid.append(row)
    return grid


def main():
    wb = load()
    out = {}
    for name in ["M", "H", "VD", "TS", "A", "V"]:
        if name in wb.sheetnames:
            ws = wb[name]
            out[name] = {
                "dims": [ws.max_row, ws.max_column],
                "sample": sample_sheet(ws, 30, 10),
            }
    # sheet with special char - index 4
    for name in wb.sheetnames:
        if name not in out and name != "G":
            ws = wb[name]
            out[name] = {"dims": [ws.max_row, ws.max_column], "sample": sample_sheet(ws, 25, 8)}
    wb.close()
    print(json.dumps(out, ensure_ascii=False, indent=2))


if __name__ == "__main__":
    main()
