"""Full Excel audit: every H row, branch logic, A inputs, sheet refs."""
import json
from pathlib import Path

from openpyxl.utils import get_column_letter

from excel_io import load_workbook, num

OUT = Path(__file__).resolve().parent.parent / "docs" / "excel_audit_full.json"


def cell_str(v):
    if v is None:
        return None
    s = str(v)
    return s if s.startswith("=") else s


def audit_h_column(h_f, h_v, col: str, label: str):
    rows = []
    for r in range(1, 185):
        val = num(h_v[f"{col}{r}"].value)
        formula = cell_str(h_f[f"{col}{r}"].value)
        cm = num(h_v[f"CM{r}"].value) if r >= 68 else None
        if val is None and formula is None and cm is None:
            continue
        rows.append({
            "row": r,
            "cm": cm,
            "value": val,
            "formula": formula,
        })
    return {"column": col, "label": label, "rows": rows}


def audit_sheet_inputs(wb_v, sheet: str, addrs: list[str]):
    ws = wb_v[sheet]
    return {a: ws[a].value for a in addrs}


def audit_m_hemşire(wb_v):
    m = wb_v["M"]
    hits = []
    for row in m.iter_rows(min_row=1, max_row=m.max_row, max_col=20):
        vals = [c.value for c in row]
        text = " ".join(str(v) for v in vals if v)
        if "HEMŞİRE" in text or "HEMSIRE" in text.upper():
            hits.append([str(v) for v in vals[:12]])
    return hits[:20]


def brut_components(h_v, col: str):
    """Replicate CO132 branch logic from cached values."""
    def gv(r):
        return num(h_v[f"{col}{r}"].value) or 0.0

    if gv(105) > 0:
        base = sum(gv(r) for r in range(105, 110))
        branch = "105-109"
    elif gv(113) > 0:
        base = sum(gv(r) for r in range(113, 118))
        branch = "113-117"
    elif gv(128) > 0:
        base = sum(gv(r) for r in range(128, 132))
        branch = "128-131"
    else:
        base = sum(gv(r) for r in range(68, 104))
        branch = "68-103"
    extras = sum(gv(r) for r in range(118, 127))
    calc = round(base + extras, 2)
    actual = gv(132)
    return {
        "branch": branch,
        "baseSum": round(base, 2),
        "extras118_126": round(extras, 2),
        "calculated132": calc,
        "actual132": actual,
        "match": abs(calc - actual) < 0.05,
        "nonzero68_103": {r: round(gv(r), 2) for r in range(68, 104) if gv(r) > 0},
        "nonzero118_126": {r: round(gv(r), 2) for r in range(118, 127) if gv(r) > 0},
        "nonzero105_131": {r: round(gv(r), 2) for r in range(105, 132) if gv(r) > 0},
    }


def main():
    wb_f = load_workbook(data_only=False)
    wb_v = load_workbook(data_only=True)
    h_v = wb_v["H"]
    h_f = wb_f["H"]

    a_addrs = [f"B{i}" for i in range(1, 25)]
    report = {
        "a_inputs": audit_sheet_inputs(wb_v, "A", a_addrs),
        "brut_cn": brut_components(h_v, "CN"),
        "brut_co": brut_components(h_v, "CO"),
        "co132_formula": cell_str(wb_f["H"]["CO132"].value),
        "co166_formula": cell_str(wb_f["H"]["CO166"].value),
        "monthly_nets_co166_177": {
            f"CO{r}": num(wb_v["H"][f"CO{r}"].value)
            for r in range(166, 178)
        },
        "monthly_nets_cn166_177": {
            f"CN{r}": num(wb_v["H"][f"CN{r}"].value)
            for r in range(166, 178)
        },
        "totals": {
            "CO178": num(wb_v["H"]["CO178"].value),
            "CO179": num(wb_v["H"]["CO179"].value),
        },
        "m_hemşire_rows": audit_m_hemşire(wb_v),
        "h_cm_nonzero": [],
        "h_co_formulas_key": {},
    }

    for r in range(68, 132):
        cm = num(h_v[f"CM{r}"].value)
        co = num(h_v[f"CO{r}"].value)
        if (cm and abs(cm) > 0) or (co and abs(co) > 0.01):
            report["h_cm_nonzero"].append({
                "row": r,
                "label": h_f.cell(r, 2).value,
                "cm": cm,
                "cn": num(h_v[f"CN{r}"].value),
                "co": co,
                "co_formula": cell_str(h_f[f"CO{r}"].value),
            })

    for r in [132, 140, 141, 166, 178, 179]:
        report["h_co_formulas_key"][f"CO{r}"] = cell_str(h_f[f"CO{r}"].value)

    # Sheet list
    report["sheets"] = wb_f.sheetnames

    OUT.parent.mkdir(parents=True, exist_ok=True)
    OUT.write_text(json.dumps(report, ensure_ascii=False, indent=2), encoding="utf-8")
    print("wrote", OUT)
    print("branch CO:", report["brut_co"]["branch"])
    print("CO132:", report["brut_co"]["actual132"], "calc:", report["brut_co"]["calculated132"])
    print("CO166:", report["monthly_nets_co166_177"]["CO166"])
    print("nonzero pay rows:", len(report["h_cm_nonzero"]))

    wb_f.close()
    wb_v.close()


if __name__ == "__main__":
    main()
