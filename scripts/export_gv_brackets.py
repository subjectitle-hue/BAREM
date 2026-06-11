"""Export H!CN36-48 GV brackets + CN49 damga + CO53-64 cross-column thresholds."""
import json
import sys
from pathlib import Path

from openpyxl.utils import column_index_from_string

sys.path.insert(0, str(Path(__file__).resolve().parent))
from excel_io import ASSETS, load_workbook, num

CN = column_index_from_string("CN")
CO = column_index_from_string("CO")


def main():
    wb = load_workbook(data_only=True, read_only=False)
    h = wb["H"]
    payload = {
        "sourceColumn": "CN",
        "temmuzColumn": "CO",
        "dilimLimits": [num(h.cell(r, CN).value) for r in range(36, 40)],
        "dilimBaseTax": [num(h.cell(r, CN).value) for r in range(40, 44)],
        "dilimRates": [num(h.cell(r, CN).value) for r in range(44, 49)],
        "damgaRate": num(h.cell(49, CN).value),
        "damgaMuafiyetBm": num(h.cell(50, CN).value),
        "agiMonthly": num(h.cell(52, CO).value) or 0.0,
        "deductionMode": "crossColumn",
        "crossColumnThresholds": {
            "gvFirstHalf": [num(h.cell(r, CO).value) or 0.0 for r in range(53, 59)],
            "gvSecondHalf": [num(h.cell(r, CO).value) or 0.0 for r in range(59, 65)],
            "dvMuafOcak": num(h.cell(50, CN).value),
            "dvMuafTem": num(h.cell(50, CO).value),
        },
    }
    ASSETS.mkdir(parents=True, exist_ok=True)
    out = ASSETS / "gv_brackets.json"
    out.write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")
    print("wrote", out)
    wb.close()


if __name__ == "__main__":
    main()
