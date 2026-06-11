"""Faz 1 — Export H sheet formula matrix (rows 4–184, cols BL..max)."""
import json
import sys
from pathlib import Path

from openpyxl.utils import get_column_letter

sys.path.insert(0, str(Path(__file__).resolve().parent))
from excel_io import ASSETS, cell_json, load_workbook, num

ROW_START = 4
ROW_END = 184
COL_START = 64  # BL — katsayı / hesap sütunları başlangıcı
PERIOD_COL_START = 4  # D — 1997 VERİ sütunu (Faz 6 yıllık seri)


def export_period_map(h):
    periods = []
    for c in range(PERIOD_COL_START, (h.max_column or 0) + 1):
        year = h.cell(2, c).value
        sem = h.cell(3, c).value
        if year is None and sem is None:
            continue
        y = int(year) if isinstance(year, (int, float)) else str(year).strip() if year else None
        periods.append({
            "col": get_column_letter(c),
            "colIndex": c,
            "year": y,
            "semester": str(sem).strip() if sem else None,
            "memurK": num(h.cell(4, c).value),
            "tabanK": num(h.cell(5, c).value),
            "yanK": num(h.cell(6, c).value),
            "enYuksek": num(h.cell(8, c).value),
        })
    return periods


def export_rows(hf, hv):
    rows = []
    formula_count = 0
    for r in range(ROW_START, ROW_END + 1):
        label = hf.cell(r, 2).value
        if not label:
            continue
        label = str(label).strip()
        bl_col = 64  # BL
        bl_cached = num(hv.cell(r, bl_col).value)
        bl_formula = hf.cell(r, bl_col).value
        bl_f = str(bl_formula) if bl_formula and str(bl_formula).startswith("=") else None

        cols = {}
        for c in range(COL_START, hf.max_column + 1):
            col_l = get_column_letter(c)
            fj = cell_json(hf.cell(r, c).value, hv.cell(r, c).value)
            if fj:
                cols[col_l] = fj
                if fj[0]:
                    formula_count += 1

        if not cols and bl_f is None and bl_cached is None:
            continue

        entry = {"r": r, "label": label}
        if bl_f or bl_cached is not None:
            entry["bl"] = [bl_f, bl_cached]
        if cols:
            entry["c"] = cols
        rows.append(entry)

    return rows, formula_count


def main():
    wb_f = load_workbook(data_only=False, read_only=False)
    wb_v = load_workbook(data_only=True, read_only=False)
    hf = wb_f["H"]
    hv = wb_v["H"]

    periods = export_period_map(hv)
    rows, formula_count = export_rows(hf, hv)

    payload = {
        "version": 1,
        "rowRange": [ROW_START, ROW_END],
        "colStart": get_column_letter(COL_START),
        "periods": periods,
        "rows": rows,
        "stats": {
            "periodColumns": len(periods),
            "dataRows": len(rows),
            "formulaCells": formula_count,
        },
    }

    ASSETS.mkdir(parents=True, exist_ok=True)
    out = ASSETS / "h_formulas.json"
    out.write_text(json.dumps(payload, ensure_ascii=False, separators=(",", ":")), encoding="utf-8")

    period_out = ASSETS / "h_period_map.json"
    period_out.write_text(
        json.dumps({"periods": periods, "activeColumns": ["CN", "CO"], "activeYear": 2026}, ensure_ascii=False, indent=2),
        encoding="utf-8",
    )

    print(f"wrote {out} ({out.stat().st_size // 1024} KB)")
    print(f"periods={len(periods)} rows={len(rows)} formulas={formula_count}")
    wb_f.close()
    wb_v.close()


if __name__ == "__main__":
    main()
