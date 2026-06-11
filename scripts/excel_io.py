"""Shared Excel workbook loader for export scripts."""
import io
from pathlib import Path

import msoffcrypto
import openpyxl

SOURCE = Path(r"C:\Users\MEMO\Downloads\Mehmet Gürer.xlsx")
PASSWORD = "erdal"
ASSETS = Path(__file__).resolve().parent.parent / "app" / "src" / "main" / "assets" / "engine"


def decrypt_buffer() -> io.BytesIO:
    with SOURCE.open("rb") as f:
        o = msoffcrypto.OfficeFile(f)
        o.load_key(password=PASSWORD)
        b = io.BytesIO()
        o.decrypt(b)
        b.seek(0)
    return b


def load_workbook(data_only: bool = True, read_only: bool = False):
    return openpyxl.load_workbook(
        decrypt_buffer(),
        data_only=data_only,
        read_only=read_only,
    )


def index_sheet(wb):
    for name in wb.sheetnames:
        if name in ("İ", "\u0130"):
            return wb[name]
    raise RuntimeError("İ sheet not found")


def num(v):
    if isinstance(v, (int, float)):
        return float(v)
    return None


def cell_json(formula_val, cached_val):
    """Compact cell: [formula|null, cached|null]."""
    f = None
    if formula_val is not None:
        s = str(formula_val)
        if s.startswith("="):
            f = s
        elif formula_val:
            f = s
    c = num(cached_val)
    if f is None and c is None:
        return None
    return [f, c]
