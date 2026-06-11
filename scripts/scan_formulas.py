import io
import re
from collections import Counter
from pathlib import Path

import msoffcrypto
import openpyxl

SOURCE = Path(r"C:\Users\MEMO\Downloads\Mehmet Gürer.xlsx")
PASSWORD = "erdal"
TARGET_SHEETS = ["H", "VD", "TS", "A", "İ", "\u0130"]


def load():
    with SOURCE.open("rb") as f:
        o = msoffcrypto.OfficeFile(f)
        o.load_key(password=PASSWORD)
        b = io.BytesIO()
        o.decrypt(b)
        b.seek(0)
    return openpyxl.load_workbook(b, data_only=False, read_only=True)


def scan(ws, max_r=250, max_c=50):
    formulas = []
    labels = []
    refs = Counter()
    for r in range(1, min(max_r, ws.max_row or 0) + 1):
        for c in range(1, min(max_c, ws.max_column or 0) + 1):
            cell = ws.cell(r, c)
            v = cell.value
            if v is None:
                continue
            coord = f"{openpyxl.utils.get_column_letter(c)}{r}"
            if isinstance(v, str) and v.startswith("="):
                formulas.append((coord, v[:200]))
                for m in re.findall(r"(?:'([^']+)'|([A-Za-z\u0130\u011E\u00DC\u00C7\u00D6\u015F\u0131\u011F\u00FC\u00E7\u00F6]+))!", str(v)):
                    s = m[0] or m[1]
                    if s and s != ws.title:
                        refs[s] += 1
            elif not isinstance(v, (int, float)) and len(str(v)) < 60:
                if any(k in str(v).lower() for k in ["memur", "işçi", "isci", "hizmet", "meslek", "derece", "çocuk", "maas", "maaş", "ücret", "seç"]):
                    labels.append((coord, str(v)))
    return formulas[:40], labels[:50], refs.most_common(20)


def main():
    wb = load()
    print("sheets:", wb.sheetnames)
    for name in wb.sheetnames:
        if name == "G" or name == "V" or name == "M":
            continue
        ws = wb[name]
        print("\n===", repr(name), "dims", ws.max_row, ws.max_column, "===")
        f, l, refs = scan(ws)
        print("labels:", len(l))
        for x in l[:25]:
            print(" ", x)
        print("formulas:", len(f))
        for x in f[:15]:
            print(" ", x)
        print("refs:", refs)
    wb.close()


if __name__ == "__main__":
    main()
