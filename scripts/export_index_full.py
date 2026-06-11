"""Faz 1 — Full İ index export (all columns E..AK, rows 7..175)."""
import json
import sys
from pathlib import Path

from openpyxl.utils import get_column_letter

sys.path.insert(0, str(Path(__file__).resolve().parent))
from excel_io import ASSETS, index_sheet, load_workbook, num

ROW_START = 7
ROW_END = 175
COL_START = 5
COL_END = 37  # AK


def period_key(year, period):
    if year is None:
        return None
    if isinstance(year, (int, float)):
        y = str(int(year))
    else:
        y = str(year).strip()
    if period is None:
        return y
    if isinstance(period, (int, float)):
        p = str(period)
    else:
        p = str(period).strip()
    return f"{y}-{p}" if p else y


def main():
    wb = load_workbook(data_only=True, read_only=False)
    ws = index_sheet(wb)

    headers = {}
    for c in range(COL_START, COL_END + 1):
        year = ws.cell(7, c).value
        period = ws.cell(8, c).value
        pk = period_key(year, period)
        if pk:
            headers[get_column_letter(c)] = {
                "col": c,
                "key": pk,
                "year": int(year) if isinstance(year, (int, float)) else year,
                "period": period,
            }

    entries = []
    for r in range(ROW_START, ROW_END + 1):
        label = ws.cell(r, 2).value
        if not label:
            continue
        label = str(label).strip()
        values = {}
        for col_l, meta in headers.items():
            v = ws.cell(r, meta["col"]).value
            if v is None:
                continue
            if isinstance(v, (int, float)):
                values[meta["key"]] = float(v)
            else:
                values[meta["key"]] = str(v).strip()
        if values:
            entries.append({"label": label, "values": values})

    payload = {
        "headers": [h["key"] for h in headers.values()],
        "headerMeta": list(headers.values()),
        "entries": entries,
    }

    ASSETS.mkdir(parents=True, exist_ok=True)
    out = ASSETS / "index_table_full.json"
    out.write_text(json.dumps(payload, ensure_ascii=False, separators=(",", ":")), encoding="utf-8")
    # keep legacy index_table.json in sync via engine bundle
    legacy = {"headers": payload["headers"], "entries": entries}
    (ASSETS / "index_table.json").write_text(
        json.dumps(legacy, ensure_ascii=False, separators=(",", ":")),
        encoding="utf-8",
    )
    print(f"wrote {out} entries={len(entries)} headers={len(headers)}")
    wb.close()


if __name__ == "__main__":
    main()
