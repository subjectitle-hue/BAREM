"""Validate Excel reference + Python engine mirror (2026 CN/CO cross-column)."""
import io
import json
import sys
from pathlib import Path

import msoffcrypto
import openpyxl
from openpyxl.utils import column_index_from_string

ROOT = Path(__file__).resolve().parent.parent
GOLDEN = ROOT / "docs" / "golden_scenarios.json"
EXCEL = Path(r"C:\Users\MEMO\Downloads\Mehmet Gürer.xlsx")
PASSWORD = "erdal"

sys.path.insert(0, str(ROOT / "scripts"))
from bl_engine import compute_scenario  # noqa: E402
from engine_calc import compute_cross_column_net, load_brackets, r2  # noqa: E402


def load_h():
    with EXCEL.open("rb") as f:
        o = msoffcrypto.OfficeFile(f)
        o.load_key(password=PASSWORD)
        b = io.BytesIO()
        o.decrypt(b)
        b.seek(0)
    return openpyxl.load_workbook(b, data_only=True, read_only=True)["H"]


def read_excel_reference(h, ocak_col, temmuz_col):
    cn = column_index_from_string(ocak_col)
    co = column_index_from_string(temmuz_col)
    return {
        "cn132": h.cell(132, cn).value,
        "co132": h.cell(132, co).value,
        "monthlyNetCo": [h.cell(166 + i, co).value for i in range(12)],
        "co178": h.cell(178, co).value,
        "co179": h.cell(179, co).value,
    }


def assert_close(label, got, exp, tol):
    if got is None or exp is None:
        print(f"  FAIL {label}: got={got} exp={exp}")
        return False
    if abs(float(got) - float(exp)) > tol:
        print(f"  FAIL {label}: got={got} exp={exp} (tol={tol})")
        return False
    print(f"  OK   {label}: {got}")
    return True


def test_engine_mirror(h, ocak_col, temmuz_col, brackets):
    cn = column_index_from_string(ocak_col)
    co = column_index_from_string(temmuz_col)
    brut_o = h.cell(132, cn).value
    brut_t = h.cell(132, co).value
    prim_o = r2(h.cell(136, cn).value)
    prim_t = r2(h.cell(136, co).value)
    matrah_o = r2(h.cell(140, cn).value)
    matrah_t = r2(h.cell(140, co).value)
    nets = compute_cross_column_net(
        brut_o, brut_t, prim_o, prim_t, matrah_o, matrah_t, brackets
    )
    ok = True
    for i, (calc, exp) in enumerate(zip(nets, [h.cell(166 + i, co).value for i in range(12)])):
        ok &= assert_close(f"engine net month {i+1}", calc, exp, 0.02)
    return ok


def validate_excel_scenario(sc, excel_ref):
    print(f"\n=== {sc['id']}: {sc['description']} ===")
    exp = sc["expected"]
    tol = exp.get("tolerance", 0.02)
    ok = True
    ok &= assert_close("cn132", excel_ref["cn132"], exp["cn132"], tol)
    ok &= assert_close("co132", excel_ref["co132"], exp["co132"], tol)
    ok &= assert_close("co178", excel_ref["co178"], exp["co178"], tol)
    ok &= assert_close("co179", excel_ref["co179"], exp["co179"], tol)
    for i, (got, want) in enumerate(zip(excel_ref["monthlyNetCo"], exp["monthlyNetCo"])):
        ok &= assert_close(f"net month {i+1}", got, want, tol)
    return ok


def validate_engine_scenario(sc, brackets):
    print(f"\n=== {sc['id']}: {sc['description']} ===")
    exp = sc["expected"]
    tol = exp.get("tolerance", 0.02)
    brut_tol = exp.get("brutTolerance", tol)
    calc = compute_scenario(sc["engineInputs"], brackets)

    ok = True
    ok &= assert_close("cn132 (engine)", calc["cn132"], exp["cn132"], brut_tol)
    ok &= assert_close("co132 (engine)", calc["co132"], exp["co132"], brut_tol)
    if "bl72" in exp:
        ok &= assert_close("BL72 tazminat %", calc["bl72"], exp["bl72"], 0.01)
    if "co178" in exp:
        ok &= assert_close("co178 (engine)", calc["co178"], exp["co178"], tol)
    if exp.get("checkNet", True):
        ok &= assert_close("co179 (engine)", calc["co179"], exp["co179"], tol)
        for i, (got, want) in enumerate(zip(calc["monthlyNetCo"], exp["monthlyNetCo"])):
            ok &= assert_close(f"net month {i+1} (engine)", got, want, tol)
    return ok


def main():
    golden = json.loads(GOLDEN.read_text(encoding="utf-8"))
    h = load_h()
    ocak = golden.get("ocakColumn", "CN")
    tem = golden.get("temmuzColumn", "CO")
    excel_ref = read_excel_reference(h, ocak, tem)
    brackets = load_brackets()

    ok = True
    for sc in golden["scenarios"]:
        mode = sc.get("validateMode", "excel")
        if mode == "engine":
            ok &= validate_engine_scenario(sc, brackets)
        else:
            ok &= validate_excel_scenario(sc, excel_ref)

    print("\n=== Python CN/CO cross-column motor testi (YHS Excel) ===")
    ok &= test_engine_mirror(h, ocak, tem, brackets)

    h.parent.close()
    print("\n" + ("PASS" if ok else "FAIL"))
    sys.exit(0 if ok else 1)


if __name__ == "__main__":
    main()
