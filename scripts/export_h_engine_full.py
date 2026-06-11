"""Export H coefficients (CN/CO — 2026 aktif dönem), GV meta, A yearly series."""
import json
import re
import sys
from pathlib import Path

from openpyxl.utils import column_index_from_string, get_column_letter

sys.path.insert(0, str(Path(__file__).resolve().parent))
from excel_io import ASSETS, load_workbook, num

MONTHS = [
    "OCAK", "ŞUBAT", "MART", "NİSAN", "MAYIS", "HAZİRAN",
    "TEMMUZ", "AĞUSTOS", "EYLÜL", "EKİM", "KASIM", "ARALIK",
]
OCAK_COL = "CN"
TEMMUZ_COL = "CO"
ACTIVE_YEAR = 2026


def export_h_coefficients(h):
    labels = {4: "memurK", 5: "tabanK", 6: "yanK", 8: "enYuksek"}
    semesters = {}
    for key, col_l in [("ocak", OCAK_COL), ("temmuz", TEMMUZ_COL)]:
        col = column_index_from_string(col_l)
        semesters[key] = {
            "column": col_l,
            **{labels[r]: num(h.cell(r, col).value) for r in labels},
        }
    return semesters


def export_gv_meta(h):
    cn = column_index_from_string(OCAK_COL)
    co = column_index_from_string(TEMMUZ_COL)
    meta = {
        "deductionMode": "crossColumn",
        "agiOran": num(h.cell(52, cn).value),
        "agiAylik": num(h.cell(52, co).value) or 0.0,
        "gvBracketsBm": [num(h.cell(r, cn).value) for r in range(53, 65)],
        "gvBracketsBn": [num(h.cell(r, co).value) for r in range(53, 65)],
        "sgkTipi": str(h.cell(65, co).value or "5434"),
        "primOran5434": num(h.cell(18, co).value),
        "primOran5510": num(h.cell(20, co).value),
        "crossColumnThresholds": {
            "gvFirstHalf": [num(h.cell(r, co).value) or 0.0 for r in range(53, 59)],
            "gvSecondHalf": [num(h.cell(r, co).value) or 0.0 for r in range(59, 65)],
            "dvMuafOcak": num(h.cell(50, cn).value),
            "dvMuafTem": num(h.cell(50, co).value),
        },
    }
    monthly = []
    for i, m in enumerate(MONTHS):
        r_gv = 142 + i
        r_dv = 154 + i
        monthly.append({
            "month": m,
            "gvRow": r_gv,
            "dvRow": r_dv,
            "gvCached": num(h.cell(r_gv, co).value),
            "dvCached": num(h.cell(r_dv, co).value),
            "gvBmCumulative": num(h.cell(r_gv, cn).value),
            "dvBmCumulative": num(h.cell(r_dv, cn).value),
        })
    meta["monthlyCached"] = monthly
    return meta


def export_a_time_series(a, h_formulas):
    years = []
    for c in range(5, (a.max_column or 0) + 1):
        y = a.cell(3, c).value
        if isinstance(y, (int, float)):
            years.append((c, int(y)))

    series = []
    for c, year in years:
        f = h_formulas.cell(4, c).value
        h_col = None
        if f and isinstance(f, str):
            m = re.search(r"H!([A-Z]+)\d+", f.upper())
            if m:
                h_col = m.group(1)
        monthly = {}
        for i, mname in enumerate(MONTHS):
            v = num(a.cell(4 + i, c).value)
            if v is not None:
                monthly[mname] = v
        if not monthly:
            continue
        series.append({
            "year": year,
            "hColumn": h_col,
            "monthlyNet": monthly,
            "firstHalfAvg": num(a.cell(16, c).value),
            "yearlyBrutGold": num(a.cell(28, c).value),
            "yearlyNetGold": num(a.cell(29, c).value),
            "yearlyBrutUsd": num(a.cell(30, c).value),
            "yearlyNetUsd": num(a.cell(31, c).value),
        })
    return series


def export_fx_by_year(idx):
    gold = {}
    usd = {}
    for c in range(5, 38):
        y = idx.cell(7, c).value
        if not isinstance(y, (int, float)):
            continue
        year = int(y)
        for r in range(90, 110):
            lb = str(idx.cell(r, 2).value or "")
            v = num(idx.cell(r, c).value)
            if v is None:
                continue
            if "ALT" in lb.upper() or "ÇEY" in lb.upper():
                gold[year] = v
            elif "DOL" in lb.upper():
                usd[year] = v
    return {"goldPerGram": gold, "usdRate": usd}


def export_h_meta(h):
    cn = column_index_from_string(OCAK_COL)
    co = column_index_from_string(TEMMUZ_COL)
    memur_k = num(h.cell(4, cn).value)
    return {
        "activePeriod": f"{ACTIVE_YEAR}-{memur_k}",
        "activeYear": ACTIVE_YEAR,
        "semesterLabel": "2026",
        "ocakColumn": OCAK_COL,
        "temmuzColumn": TEMMUZ_COL,
        "deductionMode": "crossColumn",
        "displayColumn": TEMMUZ_COL,
    }


def main():
    wb_v = load_workbook(True)
    wb_f = load_workbook(False)
    h_v = wb_v["H"]
    a_v = wb_v["A"]
    idx = wb_v["İ"]

    payload = {
        "meta": export_h_meta(h_v),
        "coefficients": export_h_coefficients(h_v),
        "gvMeta": export_gv_meta(h_v),
        "yearlySeries": export_a_time_series(a_v, wb_f["A"]),
        "fxByYear": export_fx_by_year(idx),
    }

    ASSETS.mkdir(parents=True, exist_ok=True)
    out_path = ASSETS / "h_engine_full.json"
    out_path.write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")
    print("wrote", out_path)
    print("active", payload["meta"])
    print("years", len(payload["yearlySeries"]))
    wb_v.close()
    wb_f.close()


if __name__ == "__main__":
    main()
